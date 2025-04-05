import json
from openpyxl import load_workbook

TEMPLATE_PATH = "template.xlsx"
EXCEL_OUTPUT_PATH = "result/product_details.xlsx"
JSON_OUTPUT_PATH = "result/product_details.json"


def save_to_excel(data):
    try:
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        cell_mapping = {
            "title": "B1",
            "color": "B2",
            "memory": "B3",
            "default_price": "B4",
            "discount_price": "B5",
            "item_code": "B6",
            "serial": "B7",
            "display_res": "B8",
            "item_photos": "B9",
            "item_spec": "B10"
        }

        for field, cell in cell_mapping.items():
            if field in data:
                ws[cell] = str(data[field])

        wb.save(EXCEL_OUTPUT_PATH)
        print(f"данные сохранены в excel")

    except Exception as e:
        print(f"ошибка при сохранении excel: {e}")


def save_to_json(data):
    try:
        with open(JSON_OUTPUT_PATH, 'w', encoding='utf-8') as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
        print(f"данные сохранены в json")
    except Exception as e:
        print(f"ошибка при сохранении json: {e}")